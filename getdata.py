from dataclasses import dataclass
import time
from bleak import BleakClient
import asyncio
from openpyxl import Workbook, load_workbook
import threading
import app

Mac = "A4:C1:38:6E:90:F1"  # 温度计的MAC
Interval=app.interval     # 记录的时间间隔 单位S
path='uploads/data.xlsx'
should_stop = False  # 新增的停止标志
bluetooth_status='未连接'

@dataclass
class Result:
    localtime:str
    temperature: float
    humidity: int
    voltage: float
    battery: int = 0

result = Result(localtime="", temperature=0.0, humidity=0, voltage=0.0, battery=0)


def write_init(path):
    wb=Workbook()
    ws=wb.active
    ws.cell(1,1,"localtime")
    ws.cell(1,2,"temperature")
    ws.cell(1,3,"humidity")
    ws.cell(1,4,"voltage")
    ws.cell(1,5,"battery")
    wb.save(path)

def write(path,result):
    try:
        wb = load_workbook(path)
        ws = wb.active
        valves = [result.localtime, result.temperature, result.humidity, result.voltage, result.battery]
        ws.append(valves)
        wb.save(path)
    except Exception as e:
        print(f"Error")


async def main(address,interval):
    global result,should_stop,bluetooth_status
    # Initialize the Excel file with headers if it does not exist
    try:
        load_workbook(path)
    except FileNotFoundError:
        write_init(path)

    client = BleakClient(address, timeout=30.0)
    await client.connect()
    print("连接成功")
    bluetooth_status = '连接成功'
    while not should_stop:
        buff = await client.read_gatt_char("ebe0ccc1-7a0a-4b0c-8a1a-6ff2997da3a6")
        try:
            temp = int.from_bytes(buff[0:2], byteorder='little', signed=True) / 100
            humidity = int.from_bytes(buff[2:3], byteorder='little')
            voltage = int.from_bytes(buff[3:5], byteorder='little') / 1000
            battery = round((voltage - 2) / (3.261 - 2) * 100, 2)
            result = Result(localtime="", temperature=1.0, humidity=1, voltage=0.0, battery=0)
            result = Result(
                localtime=time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()),
                temperature=temp,
                humidity=humidity,
                voltage=voltage,
                battery=battery
            )

            write(path, result)  # Write the result to the Excel file
            print(result)
            print("interval:",interval)
            print("UPLOAD_FOLDER:",app.UPLOAD_FOLDER)
            print("path:",path)
        except Exception as e:
            print(e)
        await asyncio.sleep(interval)

    await client.disconnect()
    bluetooth_status = '连接失败'


def restart_main(interval):
    global should_stop
    should_stop = False  # 重置停止标志
    asyncio.run(main(Mac, interval))

def stop_main():
    global should_stop
    should_stop = True  # 设置停止标志

# 定义一个全局变量来保存守护线程对象
thread = None

def start_thread(interval):
    global thread
    if thread is None or not thread.is_alive():
        thread = threading.Thread(target=restart_main, args=(interval,))
        thread.daemon = True
        thread.start()

def stop_thread():
    global thread
    if thread is not None and thread.is_alive():
        stop_main()
        thread.join()
        thread = None